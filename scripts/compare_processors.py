from __future__ import annotations

import argparse
import json
import re
import unicodedata
import urllib.request
from datetime import datetime
from pathlib import Path

import pandas as pd
from difflib import SequenceMatcher
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


COMPRAGAMER_PRODUCTS_URL = "https://static.compragamer.com/productos"
COMPRAGAMER_CPU_SUBCATEGORIES = {
    27: "compragamer_amd",
    48: "compragamer_intel",
}

REPORT_COLUMNS = {
    "estado": "Estado",
    "producto_normalizado": "Producto normalizado",
    "match_score": "Score match",
    "nb_codigo": "Codigo NB",
    "nb_nombre": "Producto NB",
    "nb_marca": "Marca NB",
    "nb_precio_ars": "Precio NB ARS",
    "nb_stock": "Stock NB",
    "cg_fuente": "Fuente CG",
    "cg_nombre": "Producto Compragamer",
    "cg_precio_ars": "Precio CG ARS",
    "cg_disponible": "Disponible CG",
    "diferencia_ars": "Diferencia ARS",
    "diferencia_pct": "Diferencia %",
    "cg_url": "URL Compragamer",
}


def strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", str(value))
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def normalize_text(value: str) -> str:
    value = strip_accents(value).upper()
    value = re.sub(r"[^A-Z0-9]+", " ", value)
    stopwords = {
        "PROCESADOR",
        "PROCESSOR",
        "AMD",
        "INTEL",
        "SOCKET",
        "LGA",
        "AM4",
        "AM5",
        "TURBO",
        "COOLER",
        "WRAITH",
        "STEALTH",
        "SPIRE",
        "INCLUYE",
        "INCLUIDO",
        "SIN",
        "CON",
        "PARA",
        "OEM",
        "BOX",
        "VIDEO",
        "RADEON",
        "VEGA",
        "GHZ",
        "MB",
        "W",
        "C",
        "S",
    }
    tokens = [token for token in value.split() if token not in stopwords]
    return " ".join(tokens)


def cpu_key(name: str, brand_hint: str = "") -> str:
    text = strip_accents(f"{brand_hint} {name}").upper()
    text = text.replace("CORE ULTRA", "COREULTRA")
    text = re.sub(r"\s+", " ", text)

    amd = re.search(r"\bRYZEN\s*([3579])?\s*([0-9]{4,5}[A-Z0-9]*)\b", text)
    if amd:
        tier = amd.group(1) or ""
        model = amd.group(2)
        return f"AMD|RYZEN|{tier}|{model}"

    core_ultra = re.search(r"\bCOREULTRA\s*([3579])\s*([0-9]{3,4}[A-Z]*)\b", text)
    if core_ultra:
        return f"INTEL|COREULTRA|{core_ultra.group(1)}|{core_ultra.group(2)}"

    intel_core = re.search(r"\bCORE\s*I([3579])\s*[- ]?\s*([0-9]{4,5}[A-Z]*)\b", text)
    if intel_core:
        return f"INTEL|COREI|{intel_core.group(1)}|{intel_core.group(2)}"

    pentium = re.search(r"\bPENTIUM(?:\s+GOLD)?\s*([A-Z]?[0-9]{4,5}[A-Z]*)\b", text)
    if pentium:
        return f"INTEL|PENTIUM||{pentium.group(1)}"

    celeron = re.search(r"\bCELERON\s*([A-Z]?[0-9]{4,5}[A-Z]*)\b", text)
    if celeron:
        return f"INTEL|CELERON||{celeron.group(1)}"

    return ""


def product_slug(name: str, product_id: int) -> str:
    slug = strip_accents(name)
    slug = re.sub(r"[^A-Za-z0-9]+", "_", slug).strip("_")
    return f"https://compragamer.com/producto/{slug}_{product_id}"


def load_distributor(csv_path: Path) -> pd.DataFrame:
    df = pd.read_csv(csv_path, sep=";", dtype=str, encoding="utf-8", engine="python")
    df = df[df["CATEGORIA"].fillna("").str.upper().eq("PROCESADORES")].copy()
    df["nb_codigo"] = df["CODIGO"].fillna("")
    df["nb_nombre"] = df["DETALLE"].fillna("")
    df["nb_marca"] = df["MARCA"].fillna("")
    df["nb_stock"] = pd.to_numeric(df["STOCK"], errors="coerce").fillna(0).astype(int)
    df["nb_precio_ars"] = pd.to_numeric(df["PRECIO PESOS CON IVA"], errors="coerce")
    df["match_key"] = df.apply(lambda row: cpu_key(row["nb_nombre"], row["nb_marca"]), axis=1)
    df["nb_normalizado"] = df["nb_nombre"].map(normalize_text)
    return df[["match_key", "nb_codigo", "nb_nombre", "nb_marca", "nb_precio_ars", "nb_stock", "nb_normalizado"]]


def load_compragamer() -> pd.DataFrame:
    with urllib.request.urlopen(COMPRAGAMER_PRODUCTS_URL, timeout=30) as response:
        data = json.load(response)

    rows = []
    for item in data:
        subcategory = item.get("id_subcategoria")
        price = item.get("precioEspecial") or 0
        if subcategory not in COMPRAGAMER_CPU_SUBCATEGORIES or price <= 0:
            continue
        name = item.get("nombre") or ""
        product_id = item.get("id_producto")
        rows.append(
            {
                "match_key": cpu_key(name),
                "cg_id": product_id,
                "cg_fuente": COMPRAGAMER_CPU_SUBCATEGORIES[subcategory],
                "cg_nombre": name.strip(),
                "cg_precio_ars": float(price),
                "cg_disponible": bool(item.get("vendible")),
                "cg_url": product_slug(name, product_id) + f"?cate={subcategory}",
                "cg_normalizado": normalize_text(name),
            }
        )
    return pd.DataFrame(rows)


def choose_best_by_key(df: pd.DataFrame, name_column: str) -> pd.DataFrame:
    if df.empty:
        return df
    keyed = df[df["match_key"].ne("")].copy()
    missing = df[df["match_key"].eq("")].copy()
    keyed = keyed.sort_values(name_column).drop_duplicates("match_key", keep="first")
    return pd.concat([keyed, missing], ignore_index=True)


def fuzzy_score(left: str, right: str) -> int:
    return round(SequenceMatcher(None, left, right).ratio() * 100)


def build_comparison(nb: pd.DataFrame, cg: pd.DataFrame) -> pd.DataFrame:
    nb_one = choose_best_by_key(nb, "nb_nombre")
    cg_one = choose_best_by_key(cg, "cg_nombre")
    matched = nb_one.merge(cg_one, on="match_key", how="outer", suffixes=("", ""))

    matched["producto_normalizado"] = matched.apply(
        lambda row: row["match_key"] if isinstance(row.get("match_key"), str) and row["match_key"] else row.get("nb_normalizado") or row.get("cg_normalizado"),
        axis=1,
    )
    matched["estado"] = matched.apply(
        lambda row: "repetido" if pd.notna(row.get("nb_nombre")) and pd.notna(row.get("cg_nombre")) else ("solo_distribuidor" if pd.notna(row.get("nb_nombre")) else "solo_compragamer"),
        axis=1,
    )
    matched["diferencia_ars"] = matched["cg_precio_ars"] - matched["nb_precio_ars"]
    matched["diferencia_pct"] = (matched["diferencia_ars"] / matched["nb_precio_ars"]) * 100
    matched["match_score"] = matched.apply(
        lambda row: fuzzy_score(row.get("nb_normalizado", ""), row.get("cg_normalizado", ""))
        if pd.notna(row.get("nb_normalizado")) and pd.notna(row.get("cg_normalizado"))
        else None,
        axis=1,
    )

    output_columns = [
        "estado",
        "producto_normalizado",
        "match_score",
        "nb_codigo",
        "nb_nombre",
        "nb_marca",
        "nb_precio_ars",
        "nb_stock",
        "cg_fuente",
        "cg_nombre",
        "cg_precio_ars",
        "cg_disponible",
        "diferencia_ars",
        "diferencia_pct",
        "cg_url",
    ]
    result = matched[output_columns].copy()
    sort_profit = pd.to_numeric(result["diferencia_ars"], errors="coerce").fillna(float("-inf"))
    result = (
        result.assign(_sort_profit=sort_profit)
        .sort_values(["_sort_profit", "estado", "producto_normalizado"], ascending=[False, True, True], na_position="last")
        .drop(columns="_sort_profit")
    )
    return result.fillna("-")


def autosize_columns(worksheet, max_width: int = 65) -> None:
    for index, col in enumerate(worksheet.columns, start=1):
        max_len = max(len(str(cell.value or "")) for cell in col)
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 12), max_width)


def add_table(worksheet, table_name: str) -> None:
    if worksheet.max_row < 2 or worksheet.max_column < 1:
        return
    ref = f"A1:{worksheet.cell(row=worksheet.max_row, column=worksheet.max_column).coordinate}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def style_header(worksheet, fill_color: str = "1F2937") -> None:
    header_fill = PatternFill("solid", fgColor=fill_color)
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_comparison_sheet(worksheet) -> None:
    style_header(worksheet, "111827")
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    fills = {
        "repetido": PatternFill("solid", fgColor="E7F6EA"),
        "solo_distribuidor": PatternFill("solid", fgColor="FFF4D6"),
        "solo_compragamer": PatternFill("solid", fgColor="E8F1FF"),
    }
    status_font = Font(bold=True, color="111827")
    muted_font = Font(color="64748B")

    headers = {cell.value: cell.column for cell in worksheet[1]}
    status_col = headers.get("Estado")
    diff_ars_col = headers.get("Diferencia ARS")
    diff_pct_col = headers.get("Diferencia %")
    url_col = headers.get("URL Compragamer")

    for row in worksheet.iter_rows(min_row=2):
        status = row[status_col - 1].value if status_col else None
        fill = fills.get(status)
        if fill:
            row[0].fill = fill
            row[0].font = status_font
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.value == "-":
                cell.font = muted_font
        if url_col and row[url_col - 1].value != "-":
            row[url_col - 1].hyperlink = row[url_col - 1].value
            row[url_col - 1].style = "Hyperlink"

    for column_name in ["Precio NB ARS", "Precio CG ARS", "Diferencia ARS"]:
        column = headers.get(column_name)
        if column:
            for cell in worksheet.iter_cols(min_col=column, max_col=column, min_row=2, max_row=worksheet.max_row):
                for item in cell:
                    item.number_format = '"$"#,##0'

    if diff_pct_col:
        for cell in worksheet.iter_cols(min_col=diff_pct_col, max_col=diff_pct_col, min_row=2, max_row=worksheet.max_row):
            for item in cell:
                item.number_format = '0.0"%"'

    if diff_ars_col:
        letter = worksheet.cell(row=1, column=diff_ars_col).column_letter
        data_range = f"{letter}2:{letter}{worksheet.max_row}"
        worksheet.conditional_formatting.add(
            data_range,
            CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="DFF7E7")),
        )
        worksheet.conditional_formatting.add(
            data_range,
            CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FDE2E2")),
        )

    worksheet.row_dimensions[1].height = 28
    add_table(worksheet, "TablaComparativa")
    autosize_columns(worksheet)
    worksheet.column_dimensions["A"].width = 19
    worksheet.column_dimensions["B"].width = 28
    worksheet.column_dimensions["E"].width = 48
    worksheet.column_dimensions["J"].width = 52
    worksheet.column_dimensions["O"].width = 36


def style_summary_sheet(worksheet) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet["A1"].font = Font(size=16, bold=True, color="111827")
    worksheet["A1"].fill = PatternFill("solid", fgColor="F3F4F6")
    worksheet["A1"].alignment = Alignment(vertical="center")
    worksheet.merge_cells("A1:B1")
    worksheet.row_dimensions[1].height = 30
    header_fill = PatternFill("solid", fgColor="374151")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    autosize_columns(worksheet, 40)


def style_basic_sheet(worksheet, table_name: str) -> None:
    style_header(worksheet, "374151")
    worksheet.freeze_panes = "A2"
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    add_table(worksheet, table_name)
    autosize_columns(worksheet)


def write_report(result: pd.DataFrame, nb: pd.DataFrame, cg: pd.DataFrame, output_dir: Path) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = output_dir / f"comparativa_procesadores_{timestamp}.csv"
    xlsx_path = output_dir / f"comparativa_procesadores_{timestamp}.xlsx"

    result.to_csv(csv_path, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        counts = result["estado"].value_counts().to_dict()
        summary = pd.DataFrame(
            [
                ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("Procesadores NB", len(nb)),
                ("Procesadores Compragamer AMD/Intel", len(cg)),
                ("Repetidos", counts.get("repetido", 0)),
                ("Solo distribuidor", counts.get("solo_distribuidor", 0)),
                ("Solo Compragamer", counts.get("solo_compragamer", 0)),
            ],
            columns=["Metrica", "Valor"],
        )
        summary.to_excel(writer, sheet_name="resumen", index=False, startrow=2)
        writer.sheets["resumen"]["A1"] = "Comparativa de procesadores"

        result.rename(columns=REPORT_COLUMNS).to_excel(writer, sheet_name="comparativa", index=False)
        nb.to_excel(writer, sheet_name="distribuidor_nb", index=False)
        cg.to_excel(writer, sheet_name="compragamer", index=False)

        style_summary_sheet(writer.sheets["resumen"])
        style_comparison_sheet(writer.sheets["comparativa"])
        style_basic_sheet(writer.sheets["distribuidor_nb"], "TablaDistribuidorNB")
        style_basic_sheet(writer.sheets["compragamer"], "TablaCompragamer")

    return csv_path, xlsx_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Compara procesadores NB vs Compragamer AMD/Intel.")
    parser.add_argument("--csv", required=True, type=Path, help="Ruta al CSV de lista de precios NB.")
    parser.add_argument("--out", default=Path("reports"), type=Path, help="Directorio de reportes.")
    args = parser.parse_args()

    nb = load_distributor(args.csv)
    cg = load_compragamer()
    result = build_comparison(nb, cg)
    csv_path, xlsx_path = write_report(result, nb, cg, args.out)

    counts = result["estado"].value_counts().to_dict()
    print(f"Distribuidor NB: {len(nb)} procesadores")
    print(f"Compragamer AMD/Intel: {len(cg)} procesadores")
    print(f"Repetidos: {counts.get('repetido', 0)}")
    print(f"Solo distribuidor: {counts.get('solo_distribuidor', 0)}")
    print(f"Solo Compragamer: {counts.get('solo_compragamer', 0)}")
    print(f"CSV: {csv_path}")
    print(f"Excel: {xlsx_path}")


if __name__ == "__main__":
    main()
